from scripts.paths import get_path
import json
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import cast
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell

from scripts.contracts import (
    TEMP_OUTPUT_DIR,
    ACADEMIC_JSON_FILE,
    DESTINATION_DIR,
    CourseCategory,
    FILE_PREFIX
)
def autosize_columns(ws):
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max_length + 2  # padding
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

def generate_excel_courses_list(view_type="courses_list"):

    # ----------------------------
    # 1. Validate JSON (NO TRY)
    # ----------------------------
    json_path = get_path(TEMP_OUTPUT_DIR, ACADEMIC_JSON_FILE)

    if not json_path.exists():
        raise FileNotFoundError(
            f"Stage 8 XLSX: Academic JSON not found at {json_path}"
        )

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if "courses" not in data:
        raise ValueError("Stage 8 XLSX: Missing 'courses' key in academic JSON.")

    courses_by_category = data["courses"]

    if not isinstance(courses_by_category, dict):
        raise TypeError("Stage 8 XLSX: 'courses' must be a dictionary.")

    if not courses_by_category:
        raise RuntimeError("Stage 8 XLSX: 'courses' dictionary is empty.")

    valid_categories = {cat.code for cat in CourseCategory}

    for category in courses_by_category.keys():
        if category not in valid_categories:
            raise ValueError(
                f"Stage 8 XLSX: Invalid category '{category}'."
            )

    # ----------------------------
    # 2. Excel Generation (TRY)
    # ----------------------------
    try:
        wb = Workbook()
        ws_raw = wb.active
        if ws_raw is None:
            raise RuntimeError("Workbook has no active worksheet.")
        ws = cast(Worksheet, ws_raw)
        ws.title = "Courses"

        headers = [
            "Course Code",
            "Course Title",
            "Course Category",
            "L",
            "T",
            "P",
            "X",
            "C"
        ]

        ws.append(headers)
        bold_font = Font(bold=True)

        for cell in ws[1]:  # First row
            c = cast(Cell, cell)
            c.font = bold_font
            c.alignment = Alignment(vertical="center")

        total_count = 0

        for cat in CourseCategory:
            courses = courses_by_category.get(cat.code, [])

            if not isinstance(courses, list):
                raise TypeError(
                    f"Stage 8 XLSX: Category '{cat.code}' must contain a list."
                )

            for course in courses:

                if not isinstance(course, dict):
                    raise TypeError(
                        f"Stage 8 XLSX: Course in '{cat.code}' must be a dictionary."
                    )

                meta = course.get("course_meta", {})
                if not isinstance(meta, dict):
                    raise TypeError(
                        f"Stage 8 XLSX: 'course_meta' must be a dictionary."
                    )

                course_code = course.get("course_code", "")
                if not course_code:
                    raise ValueError(
                        f"Stage 8 XLSX: Course missing 'course_code' in '{cat.code}'."
                    )

                ws.append([
                    course_code,
                    meta.get("course_title", ""),
                    meta.get("course_category", ""),
                    int(meta.get("l") or 0),
                    int(meta.get("t") or 0),
                    int(meta.get("p") or 0),
                    int(meta.get("x") or 0),
                    float(meta.get("c") or 0.0),
                ])

                total_count += 1

        if total_count == 0:
            raise RuntimeError(
                "Stage 8 XLSX: Categories exist but contain zero courses."
            )

        # Formatting
        ws.freeze_panes = "A2"

        center_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin")
        full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        center_columns = {"Course Category", "L", "T", "P", "X", "C"}

        for col_idx, header_cell in enumerate(ws[1], start=1):
            if header_cell.value in center_columns:
                for row in ws.iter_rows(
                    min_row=1,
                    min_col=col_idx,
                    max_col=col_idx
                ):
                    row[0].alignment = center_align

        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column
        ):
            for cell in row:
                cell.border = full_border

        autosize_columns(ws)

        out_path = get_path(DESTINATION_DIR, f"{FILE_PREFIX}_{view_type}.xlsx")
        out_path.parent.mkdir(parents=True, exist_ok=True)

        wb.save(out_path)

        print(f"✅ Excel file generated: {out_path}")

    except Exception as e:
        raise RuntimeError(
            f"Stage 8 XLSX: Excel generation failed: {e}"
        ) from e